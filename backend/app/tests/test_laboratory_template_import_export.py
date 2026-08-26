"""Integration tests for Laboratory Template Import/Export (bulk Excel
maintenance). See `app/services/laboratory_template_import_export.py`'s
module docstring for the two-sheet workbook format these exercise."""

import io
import uuid

import pytest
from httpx import AsyncClient
from openpyxl import Workbook, load_workbook
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.security import hash_password
from app.models.role import Role
from app.services.laboratory_template_import_export import (
    VALID_RESULT_TYPES,
    _normalize_result_type,
)

pytestmark = pytest.mark.asyncio


async def test_normalize_result_type_is_case_insensitive_but_only_for_real_values() -> None:
    """Pure-function unit test for the normalization helper itself, isolated
    from the DB/HTTP integration tests below."""
    for canonical in VALID_RESULT_TYPES:
        assert _normalize_result_type(canonical) == canonical
        assert _normalize_result_type(canonical.lower()) == canonical
        assert _normalize_result_type(canonical.upper()) == canonical
    assert _normalize_result_type("  Numeric  ") == "Numeric"
    assert _normalize_result_type("numerik") is None
    assert _normalize_result_type("") is None

XLSX_MEDIA_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


@pytest.fixture(autouse=True)
def _reset_login_rate_limit():
    from app.core.rate_limit import _memory_buckets

    _memory_buckets.clear()
    yield
    _memory_buckets.clear()


async def _login(client: AsyncClient, email: str, password: str) -> str:
    response = await client.post(
        "/api/v1/auth/login", json={"email_or_username": email, "password": password}
    )
    assert response.status_code == 200, response.text
    return response.json()["access_token"]


async def _owner_headers(client: AsyncClient, make_clinic_with_owner):
    clinic, owner, password = await make_clinic_with_owner()
    token = await _login(client, owner.email, password)
    return clinic, owner, {"Authorization": f"Bearer {token}"}


async def _make_role_login(
    db_session: AsyncSession, *, clinic_id, role_name: str, password: str = "TestPass123!"
):
    from app.models.user import User

    role = (await db_session.execute(select(Role).where(Role.name == role_name))).scalar_one()
    suffix = uuid.uuid4().hex[:8]
    email = f"{role_name.lower()}-{suffix}@example.com"
    user = User(
        clinic_id=clinic_id,
        email=email,
        username=f"{role_name.lower()}{suffix}",
        hashed_password=hash_password(password),
        first_name="Test",
        last_name=role_name,
        role_id=role.id,
        is_active=True,
    )
    db_session.add(user)
    await db_session.commit()
    await db_session.refresh(user)
    return email, user


def _workbook_bytes(template_rows: list[list], parameter_rows: list[list]) -> bytes:
    wb = Workbook()
    ws_t = wb.active
    ws_t.title = "Templates"
    ws_t.append(["ID", "Test Name", "Category", "Specimen", "Price", "Turnaround Hours", "Active"])
    for row in template_rows:
        ws_t.append(row)
    ws_p = wb.create_sheet("Parameters")
    ws_p.append(
        [
            "Template ID",
            "Template Test Name",
            "Parameter ID",
            "Parameter Name",
            "Unit",
            "Normal Range",
            "Result Type",
            "Display Order",
            "Range Low",
            "Range High",
            "Expected Normal Text",
            "Options",
            "Requires Site",
            "Section",
        ]
    )
    for row in parameter_rows:
        ws_p.append(row)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _upload(
    client: AsyncClient, path: str, headers: dict, content: bytes, filename: str = "import.xlsx"
):
    return client.post(path, headers=headers, files={"file": (filename, content, XLSX_MEDIA_TYPE)})


async def _create_template(
    client: AsyncClient,
    headers: dict,
    *,
    test_name="CBC, PLATELET",
    price="250.00",
    parameters=None,
) -> dict:
    resp = await client.post(
        "/api/v1/laboratory/templates",
        headers=headers,
        json={
            "test_name": test_name,
            "specimen_type": "Whole Blood",
            "default_price": price,
            "parameters": parameters or [],
        },
    )
    assert resp.status_code == 201, resp.text
    return resp.json()


# --- 1/2/3/4: Export ---


async def test_export_creates_valid_xlsx_with_all_templates_and_parameters(
    client: AsyncClient, make_clinic_with_owner
) -> None:
    clinic, _owner, headers = await _owner_headers(client, make_clinic_with_owner)
    await _create_template(
        client,
        headers,
        test_name="CBC, PLATELET",
        parameters=[{"parameter_name": "Hemoglobin", "unit": "g/dL", "result_type": "Numeric"}],
    )
    await _create_template(client, headers, test_name="Urinalysis", parameters=[])

    resp = await client.get("/api/v1/laboratory/templates/export", headers=headers)
    assert resp.status_code == 200
    assert resp.headers["content-type"] == XLSX_MEDIA_TYPE
    assert "laboratory-templates-" in resp.headers["content-disposition"]

    wb = load_workbook(io.BytesIO(resp.content))
    assert wb.sheetnames == ["Templates", "Parameters"]
    t_rows = list(wb["Templates"].iter_rows(values_only=True))
    p_rows = list(wb["Parameters"].iter_rows(values_only=True))
    assert len(t_rows) == 3  # header + 2 templates
    test_names = {row[1] for row in t_rows[1:]}
    assert test_names == {"CBC, PLATELET", "Urinalysis"}
    assert len(p_rows) == 2  # header + 1 parameter
    assert p_rows[1][3] == "Hemoglobin"


# --- 4: clinic isolation on export ---


async def test_export_is_clinic_isolated(client: AsyncClient, make_clinic_with_owner) -> None:
    _clinic_a, _owner_a, headers_a = await _owner_headers(client, make_clinic_with_owner)
    _clinic_b, _owner_b, headers_b = await _owner_headers(client, make_clinic_with_owner)
    await _create_template(client, headers_a, test_name="Clinic A Test")
    await _create_template(client, headers_b, test_name="Clinic B Test")

    resp = await client.get("/api/v1/laboratory/templates/export", headers=headers_a)
    wb = load_workbook(io.BytesIO(resp.content))
    names = {row[1] for row in wb["Templates"].iter_rows(values_only=True, min_row=2)}
    assert names == {"Clinic A Test"}


# --- 5/6: Import creates a new template + all its parameters ---


async def test_import_valid_workbook_creates_new_template_with_parameters(
    client: AsyncClient, make_clinic_with_owner
) -> None:
    _clinic, _owner, headers = await _owner_headers(client, make_clinic_with_owner)
    content = _workbook_bytes(
        template_rows=[["", "Urinalysis", "Clinical Microscopy", "Urine", 100, 24, "TRUE"]],
        parameter_rows=[
            [
                "",
                "Urinalysis",
                "",
                "Color",
                "",
                "",
                "Categorical",
                0,
                "",
                "",
                "",
                "Yellow, Amber",
                "FALSE",
                "",
            ],
            ["", "Urinalysis", "", "pH", "", "4.5-8", "Numeric", 1, 4.5, 8, "", "", "FALSE", ""],
        ],
    )
    preview = await _upload(client, "/api/v1/laboratory/templates/import/preview", headers, content)
    assert preview.status_code == 200, preview.text
    body = preview.json()
    assert body["template_count"] == 1
    assert body["parameter_count"] == 2
    assert body["new_template_count"] == 1
    assert body["errors"] == []
    assert body["can_commit"] is True
    assert body["diffs"][0]["action"] == "create"
    assert sorted(body["diffs"][0]["parameters"]["added"]) == ["Color", "pH"]

    commit = await _upload(client, "/api/v1/laboratory/templates/import/commit", headers, content)
    assert commit.status_code == 200, commit.text
    result = commit.json()
    assert result["created_template_count"] == 1
    assert result["parameter_count"] == 2

    listed = (await client.get("/api/v1/laboratory/templates", headers=headers)).json()
    created = next(t for t in listed if t["test_name"] == "Urinalysis")
    assert len(created["parameters"]) == 2
    assert {p["parameter_name"] for p in created["parameters"]} == {"Color", "pH"}


# --- 7/8/9: Import updates an existing template, updates existing parameters, adds new ones ---


async def test_import_updates_existing_template_and_synchronizes_parameters(
    client: AsyncClient, make_clinic_with_owner
) -> None:
    _clinic, _owner, headers = await _owner_headers(client, make_clinic_with_owner)
    template = await _create_template(
        client,
        headers,
        test_name="CBC, PLATELET",
        parameters=[
            {
                "parameter_name": "Hemoglobin",
                "unit": "g/dL",
                "normal_range": "120-160",
                "result_type": "Numeric",
            },
            {"parameter_name": "Old Parameter X", "result_type": "Text"},
        ],
    )
    template_id = template["id"]
    hgb_id = next(p["id"] for p in template["parameters"] if p["parameter_name"] == "Hemoglobin")

    content = _workbook_bytes(
        template_rows=[
            [template_id, "CBC, PLATELET (Updated)", "Hematology", "Whole Blood", 300, 12, "TRUE"]
        ],
        parameter_rows=[
            # Existing parameter, changed unit -> "changed"
            [
                template_id,
                "CBC, PLATELET (Updated)",
                hgb_id,
                "Hemoglobin",
                "g/L",
                "120-160",
                "Numeric",
                0,
                "",
                "",
                "",
                "",
                "FALSE",
                "",
            ],
            # Brand-new parameter -> "added"
            [
                template_id,
                "CBC, PLATELET (Updated)",
                "",
                "Hematocrit",
                "%",
                "36-48",
                "Numeric",
                1,
                36,
                48,
                "",
                "",
                "FALSE",
                "",
            ],
            # "Old Parameter X" intentionally omitted -> "removed"
        ],
    )
    preview = await _upload(client, "/api/v1/laboratory/templates/import/preview", headers, content)
    assert preview.status_code == 200, preview.text
    diff = preview.json()["diffs"][0]
    assert diff["action"] == "update"
    assert diff["parameters"]["changed"] == ["Hemoglobin"]
    assert diff["parameters"]["added"] == ["Hematocrit"]
    assert diff["parameters"]["removed"] == ["Old Parameter X"]

    commit = await _upload(client, "/api/v1/laboratory/templates/import/commit", headers, content)
    assert commit.status_code == 200, commit.text
    assert commit.json()["updated_template_count"] == 1

    refreshed = (await client.get("/api/v1/laboratory/templates", headers=headers)).json()
    updated = next(t for t in refreshed if t["id"] == template_id)
    assert updated["test_name"] == "CBC, PLATELET (Updated)"
    assert float(updated["default_price"]) == 300.0
    names = {p["parameter_name"] for p in updated["parameters"]}
    assert names == {"Hemoglobin", "Hematocrit"}
    assert "Old Parameter X" not in names
    hgb = next(p for p in updated["parameters"] if p["parameter_name"] == "Hemoglobin")
    assert hgb["unit"] == "g/L"


# --- 10: removed-parameter handling matches the app's existing (full-replace)
# template-update semantics ---


async def test_import_parameter_removal_matches_existing_update_semantics(
    client: AsyncClient, make_clinic_with_owner
) -> None:
    """The app's existing `PATCH /templates/{id}` already fully replaces a
    template's parameter list whenever `parameters` is sent (see
    `LaboratoryRepository.update_template`) - Import reuses that exact
    mechanism, so an omitted parameter is genuinely removed on commit. The
    Preview step is what prevents this from ever being silent - it lists
    every removal explicitly (`removed`) before the user confirms."""
    _clinic, _owner, headers = await _owner_headers(client, make_clinic_with_owner)
    template = await _create_template(
        client,
        headers,
        test_name="Panel",
        parameters=[
            {"parameter_name": "A", "result_type": "Text"},
            {"parameter_name": "B", "result_type": "Text"},
        ],
    )
    content = _workbook_bytes(
        template_rows=[[template["id"], "Panel", "", "", 0, "", "TRUE"]],
        parameter_rows=[
            [template["id"], "Panel", "", "A", "", "", "Text", 0, "", "", "", "", "FALSE", ""]
        ],
    )
    preview = (
        await _upload(client, "/api/v1/laboratory/templates/import/preview", headers, content)
    ).json()
    assert preview["diffs"][0]["parameters"]["removed"] == ["B"]

    commit = await _upload(client, "/api/v1/laboratory/templates/import/commit", headers, content)
    assert commit.status_code == 200
    refreshed = (await client.get("/api/v1/laboratory/templates", headers=headers)).json()
    updated = next(t for t in refreshed if t["id"] == template["id"])
    assert {p["parameter_name"] for p in updated["parameters"]} == {"A"}


# --- 11: required field rejected ---


async def test_import_rejects_missing_required_field(
    client: AsyncClient, make_clinic_with_owner
) -> None:
    _clinic, _owner, headers = await _owner_headers(client, make_clinic_with_owner)
    content = _workbook_bytes(
        template_rows=[["", "", "", "", "", "", "TRUE"]],  # missing Test Name
        parameter_rows=[],
    )
    preview = await _upload(client, "/api/v1/laboratory/templates/import/preview", headers, content)
    assert preview.status_code == 200
    body = preview.json()
    assert body["can_commit"] is False
    assert any("Test Name is required" in e["reason"] for e in body["errors"])

    commit = await _upload(client, "/api/v1/laboratory/templates/import/commit", headers, content)
    assert commit.status_code == 400


# --- 12: invalid parameter type rejected ---


async def test_import_rejects_invalid_parameter_type(
    client: AsyncClient, make_clinic_with_owner
) -> None:
    _clinic, _owner, headers = await _owner_headers(client, make_clinic_with_owner)
    content = _workbook_bytes(
        template_rows=[["", "Test A", "", "", 0, "", "TRUE"]],
        parameter_rows=[
            ["", "Test A", "", "Param", "", "", "NotARealType", 0, "", "", "", "", "FALSE", ""]
        ],
    )
    preview = await _upload(client, "/api/v1/laboratory/templates/import/preview", headers, content)
    body = preview.json()
    assert body["can_commit"] is False
    assert any("Invalid Result Type" in e["reason"] for e in body["errors"])


# --- Result Type normalization: case-insensitive input, canonical storage ---
# A hand-edited/human-typed workbook cell like "numeric" is semantically the
# same as the canonical "Numeric" the schema/model require - only genuinely
# unrecognized values (misspellings, other words) should be rejected. See
# `laboratory_template_import_export.py::_normalize_result_type`.


@pytest.mark.parametrize(
    "raw_result_type", ["numeric", "NUMERIC", "Numeric", "NuMeRiC", " numeric "]
)
async def test_import_normalizes_result_type_case_and_commits_as_canonical(
    client: AsyncClient, make_clinic_with_owner, raw_result_type: str
) -> None:
    _clinic, _owner, headers = await _owner_headers(client, make_clinic_with_owner)
    content = _workbook_bytes(
        template_rows=[["", "β-HCG (Serum)", "Immunology", "Serum", 300, 24, "TRUE"]],
        parameter_rows=[
            [
                "", "β-HCG (Serum)", "", "Result", "mIU/mL", "<5 mIU/mL", raw_result_type,
                0, "", "", "", "", "FALSE", "",
            ]
        ],
    )
    preview = await _upload(client, "/api/v1/laboratory/templates/import/preview", headers, content)
    body = preview.json()
    assert body["errors"] == [], body["errors"]
    assert body["can_commit"] is True

    commit = await _upload(client, "/api/v1/laboratory/templates/import/commit", headers, content)
    assert commit.status_code == 200, commit.text

    listed = (await client.get("/api/v1/laboratory/templates", headers=headers)).json()
    template = next(t for t in listed if t["test_name"] == "β-HCG (Serum)")
    param = template["parameters"][0]
    # Canonical enum value stored, regardless of the casing typed in the sheet.
    assert param["result_type"] == "Numeric"
    # The clinical reference value itself is untouched free text - normalization
    # only rewrites the Result Type cell, never the Normal Range content.
    assert param["normal_range"] == "<5 mIU/mL"


async def test_import_still_rejects_a_genuinely_unknown_result_type_after_normalization(
    client: AsyncClient, make_clinic_with_owner
) -> None:
    """Normalization only widens accepted CASING of the five real result
    types - it must not silently accept an unrelated/misspelled value."""
    _clinic, _owner, headers = await _owner_headers(client, make_clinic_with_owner)
    content = _workbook_bytes(
        template_rows=[["", "Test A", "", "", 0, "", "TRUE"]],
        parameter_rows=[
            ["", "Test A", "", "Param", "", "", "numerik", 0, "", "", "", "", "FALSE", ""]
        ],
    )
    preview = await _upload(client, "/api/v1/laboratory/templates/import/preview", headers, content)
    body = preview.json()
    assert body["can_commit"] is False
    assert any("Invalid Result Type" in e["reason"] for e in body["errors"])


# --- 13: duplicate parameter rejected ---


async def test_import_rejects_duplicate_parameter_in_same_template(
    client: AsyncClient, make_clinic_with_owner
) -> None:
    _clinic, _owner, headers = await _owner_headers(client, make_clinic_with_owner)
    content = _workbook_bytes(
        template_rows=[["", "Test A", "", "", 0, "", "TRUE"]],
        parameter_rows=[
            ["", "Test A", "", "Hemoglobin", "", "", "Numeric", 0, "", "", "", "", "FALSE", ""],
            ["", "Test A", "", "Hemoglobin", "", "", "Numeric", 1, "", "", "", "", "FALSE", ""],
        ],
    )
    preview = await _upload(client, "/api/v1/laboratory/templates/import/preview", headers, content)
    body = preview.json()
    assert body["can_commit"] is False
    assert any("Duplicate parameter name" in e["reason"] for e in body["errors"])


# --- 14: invalid template reference rejected ---


async def test_import_rejects_parameter_referencing_unknown_template(
    client: AsyncClient, make_clinic_with_owner
) -> None:
    _clinic, _owner, headers = await _owner_headers(client, make_clinic_with_owner)
    content = _workbook_bytes(
        template_rows=[["", "Test A", "", "", 0, "", "TRUE"]],
        parameter_rows=[
            [
                "",
                "A Completely Different Name",
                "",
                "Param",
                "",
                "",
                "Numeric",
                0,
                "",
                "",
                "",
                "",
                "FALSE",
                "",
            ]
        ],
    )
    preview = await _upload(client, "/api/v1/laboratory/templates/import/preview", headers, content)
    body = preview.json()
    assert body["can_commit"] is False
    assert any("No Templates sheet row found" in e["reason"] for e in body["errors"])


# --- 15: cross-clinic template ID rejected ---


async def test_import_rejects_cross_clinic_template_id(
    client: AsyncClient, make_clinic_with_owner
) -> None:
    _clinic_a, _owner_a, headers_a = await _owner_headers(client, make_clinic_with_owner)
    _clinic_b, _owner_b, headers_b = await _owner_headers(client, make_clinic_with_owner)
    template_b = await _create_template(client, headers_b, test_name="Clinic B's Template")

    content = _workbook_bytes(
        template_rows=[[template_b["id"], "Hijacked Name", "", "", 0, "", "TRUE"]],
        parameter_rows=[],
    )
    preview = await _upload(
        client, "/api/v1/laboratory/templates/import/preview", headers_a, content
    )
    body = preview.json()
    assert body["can_commit"] is False
    assert any("belongs to a different clinic" in e["reason"] for e in body["errors"])

    commit = await _upload(client, "/api/v1/laboratory/templates/import/commit", headers_a, content)
    assert commit.status_code == 400

    # Clinic B's template must be completely untouched.
    unchanged = (await client.get("/api/v1/laboratory/templates", headers=headers_b)).json()
    assert (
        next(t for t in unchanged if t["id"] == template_b["id"])["test_name"]
        == "Clinic B's Template"
    )


# --- 16/17: entire transaction rolls back on failure, no partial modification ---


async def test_import_transaction_rolls_back_entirely_on_a_later_row_failure(
    client: AsyncClient, make_clinic_with_owner
) -> None:
    """Template 1 is valid; Template 2 references a cross-clinic ID (must
    fail at Preview/Commit validation). Since Commit independently
    re-validates the whole file before writing anything, NEITHER template
    should be created - not even the valid one."""
    _clinic_a, _owner_a, headers_a = await _owner_headers(client, make_clinic_with_owner)
    _clinic_b, _owner_b, headers_b = await _owner_headers(client, make_clinic_with_owner)
    template_b = await _create_template(client, headers_b, test_name="Clinic B's Template")

    content = _workbook_bytes(
        template_rows=[
            ["", "Valid New Template", "", "", 100, "", "TRUE"],
            [template_b["id"], "Should Fail", "", "", 0, "", "TRUE"],
        ],
        parameter_rows=[],
    )
    commit = await _upload(client, "/api/v1/laboratory/templates/import/commit", headers_a, content)
    assert commit.status_code == 400

    listed = (await client.get("/api/v1/laboratory/templates", headers=headers_a)).json()
    assert not any(t["test_name"] == "Valid New Template" for t in listed)


# --- 18: existing manual template CRUD remains unchanged ---


async def test_manual_template_crud_unaffected_by_import_feature(
    client: AsyncClient, make_clinic_with_owner
) -> None:
    _clinic, _owner, headers = await _owner_headers(client, make_clinic_with_owner)
    created = await _create_template(client, headers, test_name="Manual Template")
    update = await client.patch(
        f"/api/v1/laboratory/templates/{created['id']}",
        headers=headers,
        json={"default_price": "99.00"},
    )
    assert update.status_code == 200
    assert float(update.json()["default_price"]) == 99.0


# --- Structural validation: missing sheet ---


async def test_import_rejects_workbook_missing_required_sheet(
    client: AsyncClient, make_clinic_with_owner
) -> None:
    _clinic, _owner, headers = await _owner_headers(client, make_clinic_with_owner)
    wb = Workbook()
    wb.active.title = "NotTemplates"
    buf = io.BytesIO()
    wb.save(buf)
    preview = await _upload(
        client, "/api/v1/laboratory/templates/import/preview", headers, buf.getvalue()
    )
    assert preview.status_code == 400


# --- Permissions: Doctor (view-only) can export but not import ---


async def test_export_allowed_for_view_role_import_forbidden(
    client: AsyncClient, make_clinic_with_owner, db_session
) -> None:
    clinic, _owner, owner_headers = await _owner_headers(client, make_clinic_with_owner)
    await _create_template(client, owner_headers, test_name="CBC, PLATELET")
    doc_email, _doc = await _make_role_login(db_session, clinic_id=clinic.id, role_name="Doctor")
    doc_token = await _login(client, doc_email, "TestPass123!")
    doc_headers = {"Authorization": f"Bearer {doc_token}"}

    export = await client.get("/api/v1/laboratory/templates/export", headers=doc_headers)
    assert export.status_code == 200

    content = _workbook_bytes(template_rows=[["", "New", "", "", 0, "", "TRUE"]], parameter_rows=[])
    preview = await _upload(
        client, "/api/v1/laboratory/templates/import/preview", doc_headers, content
    )
    assert preview.status_code == 403

    blank = await client.get("/api/v1/laboratory/templates/import/blank", headers=doc_headers)
    assert blank.status_code == 403
