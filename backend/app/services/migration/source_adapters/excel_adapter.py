"""Excel source adapter - fully implemented via `openpyxl`. One workbook,
one sheet per entity type (sheet name == entity type, e.g. "Patients",
"Doctors") - a configurable sheet-name-to-entity mapping can be layered
on top by the mapping service if a legacy export uses different sheet
names."""

import io
from collections.abc import AsyncIterator
from typing import Any

from openpyxl import load_workbook


class ExcelSourceAdapter:
    """Reads rows from named sheets of a single uploaded workbook, in
    batches. `connection_config` expects `{"file": path_or_bytes}`."""

    def __init__(self, connection_config: dict[str, Any]) -> None:
        self.connection_config = connection_config
        self._connected = False
        self._sheets: dict[str, list[dict[str, Any]]] = {}

    async def connect(self) -> None:
        source = self.connection_config.get("file")
        if isinstance(source, (bytes, bytearray)):
            wb = load_workbook(io.BytesIO(source), read_only=True, data_only=True)
        else:
            wb = load_workbook(source, read_only=True, data_only=True)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows_iter = ws.iter_rows(values_only=True)
            try:
                header = [str(c) if c is not None else "" for c in next(rows_iter)]
            except StopIteration:
                self._sheets[sheet_name] = []
                continue
            sheet_rows = []
            for raw_row in rows_iter:
                if all(v is None for v in raw_row):
                    continue
                sheet_rows.append({header[i]: raw_row[i] for i in range(len(header)) if i < len(raw_row)})
            self._sheets[sheet_name] = sheet_rows
        wb.close()
        self._connected = True

    async def analyze_schema(self) -> dict[str, list[str]]:
        return {name: (list(rows[0].keys()) if rows else []) for name, rows in self._sheets.items()}

    async def read_table(
        self, name: str, *, batch_size: int = 500, offset: int = 0
    ) -> AsyncIterator[list[dict[str, Any]]]:
        rows = self._sheets.get(name, [])
        for start in range(offset, len(rows), batch_size):
            yield rows[start : start + batch_size]

    async def count_rows(self, name: str) -> int:
        return len(self._sheets.get(name, []))

    async def close(self) -> None:
        self._sheets.clear()
        self._connected = False
