"""Laboratory Template Import/Export (bulk Excel maintenance).

Two-sheet `.xlsx` workbook - "Templates" (template-level fields) and
"Parameters" (every parameter belonging to a template, one row each), joined
within the file by `Template ID` (existing templates) or `Template Test
Name` (new templates, which have no ID yet). Mirrors exactly the same
fields `LaboratoryTemplateCreate`/`LaboratoryTemplateParameterCreate`
(`app/schemas/laboratory.py`) already accept - no invented fields.

This module is intentionally DB-free: it only reads/writes workbook bytes
and produces plain dataclasses. All validation that requires the database
(tenant ownership, "does this ID exist") lives in
`LaboratoryService.preview_import`/`commit_import`, which is the only place
that reuses the existing `LaboratoryTemplateCreate`/`Update` Pydantic
schemas for row-level field validation - see that module for why.

Parsing follows the exact same `openpyxl.load_workbook(..., read_only=True,
data_only=True)` + first-row-is-header pattern already established in
`app/services/migration/source_adapters/excel_adapter.py`, the only other
place this codebase reads an uploaded `.xlsx`.
"""

from __future__ import annotations

import io
from dataclasses import dataclass, field
from decimal import Decimal, InvalidOperation
from typing import Any
from uuid import UUID

from openpyxl import Workbook, load_workbook
from openpyxl.comments import Comment

TEMPLATES_SHEET = "Templates"
PARAMETERS_SHEET = "Parameters"

TEMPLATE_HEADERS = [
    "ID",
    "Test Name",
    "Category",
    "Specimen",
    "Price",
    "Turnaround Hours",
    "Active",
]
PARAMETER_HEADERS = [
    "Template ID",
    "Template Test Name",
    "Parameter ID",
    "Parameter Name",
    "Unit",
    "Normal Range",
    "Result Type",
    "Display Order",
    "Range Low",
    "Range High",
    "Expected Normal Text",
    "Options",
    "Requires Site",
    "Section",
]

VALID_RESULT_TYPES = {"Numeric", "Text", "Categorical", "Microscopy", "Titer"}
_TRUE_VALUES = {"true", "1", "yes", "y", "active"}
_FALSE_VALUES = {"false", "0", "no", "n", "inactive"}


# --- Row-level parse results (structural parsing only - no DB access) ---


@dataclass
class ParseIssue:
    severity: str  # "error" | "warning"
    sheet: str
    row: int
    template: str | None
    parameter: str | None
    reason: str


@dataclass
class ParsedTemplateRow:
    row: int
    raw_id: str | None
    id: UUID | None
    test_name: str
    test_category: str | None
    specimen_type: str | None
    default_price: Decimal
    turnaround_time_hours: int | None
    is_active: bool


@dataclass
class ParsedParameterRow:
    row: int
    raw_template_id: str | None
    template_id: UUID | None
    template_test_name: str
    raw_parameter_id: str | None
    parameter_id: UUID | None
    parameter_name: str
    unit: str | None
    normal_range: str | None
    result_type: str
    display_order: int | None
    range_low: Decimal | None
    range_high: Decimal | None
    expected_normal_text: str | None
    options: list[str] | None
    requires_site: bool
    section: str | None


@dataclass
class ParsedWorkbook:
    templates: list[ParsedTemplateRow] = field(default_factory=list)
    parameters: list[ParsedParameterRow] = field(default_factory=list)
    issues: list[ParseIssue] = field(default_factory=list)


class WorkbookStructureError(ValueError):
    """Raised when the uploaded file isn't a readable two-sheet workbook -
    a hard stop before any row-level validation can even begin."""


def _read_sheet_rows(content: bytes, sheet_name: str) -> list[dict[str, Any]]:
    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    try:
        if sheet_name not in wb.sheetnames:
            raise WorkbookStructureError(f'Missing required sheet "{sheet_name}".')
        ws = wb[sheet_name]
        rows_iter = ws.iter_rows(values_only=True)
        try:
            header = [str(c).strip() if c is not None else "" for c in next(rows_iter)]
        except StopIteration:
            return []
        rows = []
        for raw_row in rows_iter:
            if all(v is None for v in raw_row):
                continue
            rows.append({header[i]: raw_row[i] for i in range(len(header)) if i < len(raw_row)})
        return rows
    finally:
        wb.close()


def _s(value: Any) -> str | None:
    """Cell value to a trimmed string, or None if blank."""
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def _parse_bool(value: Any, *, default: bool) -> tuple[bool | None, bool]:
    """Returns (value, ok). `ok=False` means the cell had content that
    could not be parsed as a boolean."""
    if value is None or value == "":
        return default, True
    if isinstance(value, bool):
        return value, True
    text = str(value).strip().lower()
    if text in _TRUE_VALUES:
        return True, True
    if text in _FALSE_VALUES:
        return False, True
    return None, False


def _parse_decimal(value: Any) -> tuple[Decimal | None, bool]:
    if value is None or value == "":
        return None, True
    try:
        return Decimal(str(value).strip()), True
    except (InvalidOperation, ValueError):
        return None, False


def _parse_int(value: Any) -> tuple[int | None, bool]:
    if value is None or value == "":
        return None, True
    try:
        # Excel often stores whole numbers as floats (e.g. 24.0).
        f = float(value)
        if not f.is_integer():
            return None, False
        return int(f), True
    except (TypeError, ValueError):
        return None, False


def _parse_uuid(value: Any) -> tuple[UUID | None, bool]:
    if value is None or value == "":
        return None, True
    try:
        return UUID(str(value).strip()), True
    except ValueError:
        return None, False


def _parse_options(value: Any) -> list[str] | None:
    text = _s(value)
    if text is None:
        return None
    return [part.strip() for part in text.split(",") if part.strip()]


def parse_workbook(content: bytes) -> ParsedWorkbook:
    """Structural parse only - never touches the database. Raises
    `WorkbookStructureError` if the file isn't a readable workbook with
    both required sheets; row-level problems are collected as `issues`
    instead of raising, so the caller can still show a full preview."""
    try:
        template_rows = _read_sheet_rows(content, TEMPLATES_SHEET)
        parameter_rows = _read_sheet_rows(content, PARAMETERS_SHEET)
    except WorkbookStructureError:
        raise
    except Exception as exc:  # openpyxl raises various error types for a corrupt/non-xlsx file
        raise WorkbookStructureError(
            f"Could not read the uploaded file as an Excel workbook: {exc}"
        ) from exc

    result = ParsedWorkbook()

    for i, row in enumerate(template_rows, start=2):  # header is row 1
        test_name = _s(row.get("Test Name"))
        raw_id = _s(row.get("ID"))
        template_id, id_ok = _parse_uuid(raw_id)
        if not id_ok:
            result.issues.append(
                ParseIssue(
                    "error",
                    TEMPLATES_SHEET,
                    i,
                    test_name,
                    None,
                    f'Invalid ID "{raw_id}" - must be a UUID or blank.',
                )
            )
        if not test_name:
            result.issues.append(
                ParseIssue("error", TEMPLATES_SHEET, i, None, None, "Test Name is required.")
            )
            continue
        price, price_ok = _parse_decimal(row.get("Price"))
        if not price_ok:
            result.issues.append(
                ParseIssue(
                    "error",
                    TEMPLATES_SHEET,
                    i,
                    test_name,
                    None,
                    f'Invalid Price "{row.get("Price")}".',
                )
            )
        turnaround, turnaround_ok = _parse_int(row.get("Turnaround Hours"))
        if not turnaround_ok:
            result.issues.append(
                ParseIssue(
                    "error",
                    TEMPLATES_SHEET,
                    i,
                    test_name,
                    None,
                    f'Invalid Turnaround Hours "{row.get("Turnaround Hours")}".',
                )
            )
        is_active, active_ok = _parse_bool(row.get("Active"), default=True)
        if not active_ok:
            result.issues.append(
                ParseIssue(
                    "error",
                    TEMPLATES_SHEET,
                    i,
                    test_name,
                    None,
                    f'Invalid Active value "{row.get("Active")}" - use TRUE/FALSE.',
                )
            )

        result.templates.append(
            ParsedTemplateRow(
                row=i,
                raw_id=raw_id,
                id=template_id,
                test_name=test_name,
                test_category=_s(row.get("Category")),
                specimen_type=_s(row.get("Specimen")),
                default_price=price if price_ok and price is not None else Decimal("0"),
                turnaround_time_hours=turnaround if turnaround_ok else None,
                is_active=is_active if active_ok and is_active is not None else True,
            )
        )

    for i, row in enumerate(parameter_rows, start=2):
        template_test_name = _s(row.get("Template Test Name"))
        parameter_name = _s(row.get("Parameter Name"))
        raw_template_id = _s(row.get("Template ID"))
        template_id, tid_ok = _parse_uuid(raw_template_id)
        if not tid_ok:
            result.issues.append(
                ParseIssue(
                    "error",
                    PARAMETERS_SHEET,
                    i,
                    template_test_name,
                    parameter_name,
                    f'Invalid Template ID "{raw_template_id}".',
                )
            )
        raw_param_id = _s(row.get("Parameter ID"))
        parameter_id, pid_ok = _parse_uuid(raw_param_id)
        if not pid_ok:
            result.issues.append(
                ParseIssue(
                    "error",
                    PARAMETERS_SHEET,
                    i,
                    template_test_name,
                    parameter_name,
                    f'Invalid Parameter ID "{raw_param_id}".',
                )
            )

        if not template_test_name:
            result.issues.append(
                ParseIssue(
                    "error",
                    PARAMETERS_SHEET,
                    i,
                    None,
                    parameter_name,
                    "Template Test Name is required.",
                )
            )
            continue
        if not parameter_name:
            result.issues.append(
                ParseIssue(
                    "error",
                    PARAMETERS_SHEET,
                    i,
                    template_test_name,
                    None,
                    "Parameter Name is required.",
                )
            )
            continue

        result_type = _s(row.get("Result Type")) or "Numeric"
        if result_type not in VALID_RESULT_TYPES:
            result.issues.append(
                ParseIssue(
                    "error",
                    PARAMETERS_SHEET,
                    i,
                    template_test_name,
                    parameter_name,
                    (
                        f'Invalid Result Type "{result_type}" - must be one of '
                        f'{", ".join(sorted(VALID_RESULT_TYPES))}.'
                    ),
                )
            )

        display_order, order_ok = _parse_int(row.get("Display Order"))
        if not order_ok:
            result.issues.append(
                ParseIssue(
                    "error",
                    PARAMETERS_SHEET,
                    i,
                    template_test_name,
                    parameter_name,
                    f'Invalid Display Order "{row.get("Display Order")}".',
                )
            )
        elif display_order is not None and display_order < 0:
            result.issues.append(
                ParseIssue(
                    "error",
                    PARAMETERS_SHEET,
                    i,
                    template_test_name,
                    parameter_name,
                    "Display Order cannot be negative.",
                )
            )

        range_low, low_ok = _parse_decimal(row.get("Range Low"))
        if not low_ok:
            result.issues.append(
                ParseIssue(
                    "error",
                    PARAMETERS_SHEET,
                    i,
                    template_test_name,
                    parameter_name,
                    f'Invalid Range Low "{row.get("Range Low")}".',
                )
            )
        range_high, high_ok = _parse_decimal(row.get("Range High"))
        if not high_ok:
            result.issues.append(
                ParseIssue(
                    "error",
                    PARAMETERS_SHEET,
                    i,
                    template_test_name,
                    parameter_name,
                    f'Invalid Range High "{row.get("Range High")}".',
                )
            )

        requires_site, site_ok = _parse_bool(row.get("Requires Site"), default=False)
        if not site_ok:
            result.issues.append(
                ParseIssue(
                    "error",
                    PARAMETERS_SHEET,
                    i,
                    template_test_name,
                    parameter_name,
                    f'Invalid Requires Site value "{row.get("Requires Site")}" - use TRUE/FALSE.',
                )
            )

        options = _parse_options(row.get("Options"))
        if result_type == "Categorical" and not options:
            result.issues.append(
                ParseIssue(
                    "warning",
                    PARAMETERS_SHEET,
                    i,
                    template_test_name,
                    parameter_name,
                    "Categorical parameter has no Options configured.",
                )
            )

        result.parameters.append(
            ParsedParameterRow(
                row=i,
                raw_template_id=raw_template_id,
                template_id=template_id,
                template_test_name=template_test_name,
                raw_parameter_id=raw_param_id,
                parameter_id=parameter_id,
                parameter_name=parameter_name,
                unit=_s(row.get("Unit")),
                normal_range=_s(row.get("Normal Range")),
                result_type=result_type if result_type in VALID_RESULT_TYPES else "Numeric",
                display_order=display_order if order_ok else None,
                range_low=range_low if low_ok else None,
                range_high=range_high if high_ok else None,
                expected_normal_text=_s(row.get("Expected Normal Text")),
                options=options,
                requires_site=requires_site if site_ok and requires_site is not None else False,
                section=_s(row.get("Section")),
            )
        )

    return result


# --- Workbook generation (export + blank-template download) ---


def _write_headers(ws, headers: list[str]) -> None:
    for col, title in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=title)
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = max(
            12, len(title) + 2
        )


def build_export_workbook(templates: list[dict]) -> bytes:
    """`templates`: list of dicts each shaped like `LaboratoryTemplateRead.
    model_dump()` (id/test_name/.../parameters). Only reads - never touches
    the database itself, so the caller is responsible for tenant scoping
    (only pass this clinic's own templates)."""
    wb = Workbook()
    ws_t = wb.active
    ws_t.title = TEMPLATES_SHEET
    _write_headers(ws_t, TEMPLATE_HEADERS)
    ws_p = wb.create_sheet(PARAMETERS_SHEET)
    _write_headers(ws_p, PARAMETER_HEADERS)

    t_row = 2
    p_row = 2
    for t in templates:
        ws_t.append(
            [
                str(t["id"]),
                t["test_name"],
                t.get("test_category"),
                t.get("specimen_type"),
                float(t["default_price"]),
                t.get("turnaround_time_hours"),
                "TRUE" if t["is_active"] else "FALSE",
            ]
        )
        t_row += 1
        for p in t.get("parameters", []):
            ws_p.append(
                [
                    str(t["id"]),
                    t["test_name"],
                    str(p["id"]),
                    p["parameter_name"],
                    p.get("unit"),
                    p.get("normal_range"),
                    p["result_type"],
                    p.get("display_order", 0),
                    float(p["range_low"]) if p.get("range_low") is not None else None,
                    float(p["range_high"]) if p.get("range_high") is not None else None,
                    p.get("expected_normal_text"),
                    ", ".join(p["options"]) if p.get("options") else None,
                    "TRUE" if p.get("requires_site") else "FALSE",
                    p.get("section"),
                ]
            )
            p_row += 1

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_blank_import_workbook() -> bytes:
    """Downloadable starter workbook: headers + one illustrative example
    row per sheet. Never written to the database - purely a local file the
    user fills in and re-uploads."""
    wb = Workbook()
    ws_t = wb.active
    ws_t.title = TEMPLATES_SHEET
    _write_headers(ws_t, TEMPLATE_HEADERS)
    ws_t.cell(row=1, column=1).comment = Comment(
        "Leave blank to create a new template. Fill in an existing template's ID to update it "
        "(must belong to your clinic) instead of creating a duplicate.",
        "CONNECT.PH",
    )
    ws_t.append(["", "CBC, PLATELET", "Hematology", "Whole Blood", 250, 24, "TRUE"])

    ws_p = wb.create_sheet(PARAMETERS_SHEET)
    _write_headers(ws_p, PARAMETER_HEADERS)
    ws_p.cell(row=1, column=2).comment = Comment(
        "Must exactly match a Test Name on the Templates sheet - this is how a parameter row is "
        "matched to its template within this file.",
        "CONNECT.PH",
    )
    ws_p.append(
        [
            "",
            "CBC, PLATELET",
            "",
            "Hemoglobin",
            "g/dL",
            "120-170",
            "Numeric",
            0,
            120,
            170,
            "",
            "",
            "FALSE",
            "",
        ]
    )
    ws_p.append(
        [
            "",
            "CBC, PLATELET",
            "",
            "Hematocrit",
            "%",
            "0.36-0.48",
            "Numeric",
            1,
            0.36,
            0.48,
            "",
            "",
            "FALSE",
            "",
        ]
    )

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
